#!/usr/bin/env python3
"""Rebuild recipes.json from Meal_Generator.xlsx and tag soups in the xlsx.

A recipe is a soup if its name matches /soup|湯|汤/i. The Soup flag is written to
a 'Soup' column on the 'All Recipes' tab (added if missing) without disturbing the
RAND helper columns G-O, and copied into recipes.json as `isSoup`.

Usage:  python3 build_recipes.py
"""
import json
import re

import openpyxl

XLSX = "Meal_Generator.xlsx"
SHEET = "All Recipes"
SOUP_RE = re.compile(r"soup|湯|汤", re.I)


def tag_soups():
    """Add/refresh the Soup column in the xlsx, preserving formulas."""
    wb = openpyxl.load_workbook(XLSX)  # keep formulas (no data_only)
    ws = wb[SHEET]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    soup_col = header.index("Soup") + 1 if "Soup" in header else ws.max_column + 1
    ws.cell(1, soup_col).value = "Soup"
    n = 0
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if name and SOUP_RE.search(str(name)):
            ws.cell(r, soup_col).value = True
            n += 1
        else:
            ws.cell(r, soup_col).value = None
    wb.save(XLSX)
    return n


def build_json():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    recipes = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        name = str(name).strip()
        ing = ws.cell(r, 2).value
        ing = "" if (ing is None or str(ing).strip() == "—") else str(ing).strip()
        meals = ws.cell(r, 3).value
        recipes.append({
            "id": r - 1,
            "name": name,
            "ingredients": ing,
            "meals": [x.strip() for x in str(meals).split(",")] if meals else [],
            "cuisine": (ws.cell(r, 4).value or "").strip(),
            "confidence": (ws.cell(r, 5).value or "").strip(),
            "link": ws.cell(r, 6).value or "",
            "isSoup": bool(SOUP_RE.search(name)),
        })
    with open("recipes.json", "w", encoding="utf-8") as f:
        json.dump(recipes, f, ensure_ascii=False, indent=0)
    return recipes


if __name__ == "__main__":
    soups = tag_soups()
    recipes = build_json()
    print(f"Tagged {soups} soups · wrote {len(recipes)} recipes to recipes.json")
